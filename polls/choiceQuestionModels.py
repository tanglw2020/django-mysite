from django.db import models
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from django.utils.html import format_html, format_html_join
from django.utils import timezone
from polls.fileModels import *
from openpyxl import load_workbook
from openpyxl import cell

ANSWER_CHOICES = [
    ('1','1'),
    ('2','2'),
    ('3','3'),
    ('4','4'),
]

class ChoiceQuestion(models.Model):

    class Meta:
        verbose_name = '题目-选择题'
        verbose_name_plural = '题目-选择题'

    def __str__(self):
        return '选择题'+str(self.id)

    def description(self):
        choice_list = [
            ['white', 'A. '+ self.choice_1], 
            ['white', 'B. '+  self.choice_2], 
            ['white', 'C. '+  self.choice_3], 
            ['white', 'D. '+  self.choice_4], 
            ]
        choice_list[int(self.answer)-1][0] = 'Lime'

        return format_html("<ul>") + \
                format_html_join(
                '\n', '<li style="background-color:{};">{}</li>',
                ((x[0], x[1]) for x in choice_list)
                ) \
                + format_html("</ul>")
    description.short_description = '题目内容'

    question_text = models.TextField('题干')
    choice_1 = models.CharField('选项1', max_length=200,  default='')
    choice_2 = models.CharField('选项2', max_length=200,  default='')
    choice_3 = models.CharField('选项3', max_length=200,  default='')
    choice_4 = models.CharField('选项4', max_length=200,  default='')

    answer = models.CharField('答案', max_length=2, choices= ANSWER_CHOICES)


class ChoiceQuestionImporter(models.Model):
    class Meta:
        verbose_name = '题目-批量导入选择题'
        verbose_name_plural = '题目-批量导入选择题'

    def __str__(self):
        return '导入选择题'+str(self.id)

    pub_date = models.DateTimeField('导入时间', null=True, default=timezone.now)
    upload_description_file = models.FileField(upload_to='upload_import_files/', null=True, blank=True, 
    validators=[validate_xlsx], verbose_name='上传选择文件[.xlsx]')

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs) 
        print(self.upload_description_file, 'saved')

        wb = load_workbook(self.upload_description_file.path)
        ws = wb.active 
        print('sheetnames: ', wb.sheetnames)
        print('ws1: ', 
        ws['B1'].value, 
        ws['C1'].value, 
        ws['D1'].value, 
        ws['E1'].value,  
        ws['F1'].value,  
        ws['G1'].value,)
        if  ws['B1'].value != '题干' or\
        ws['C1'].value != '选项1' or\
        ws['D1'].value != '选项2' or\
        ws['E1'].value != '选项3' or\
        ws['F1'].value != '选项4' or\
        ws['G1'].value != '正确选项':
            print('导入错误：文件内容格式不对！')
            return

        prep = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        row_id = 1
        empty_rows = 0
        while empty_rows<10:
            row_id = row_id + 1
            question_text = ws[prep[0]+str(row_id)].value
            if question_text is None or question_text.strip() == '':
                empty_rows = empty_rows +1
                continue
            
            empty_rows = 0

            choice_1 = ws[prep[1]+str(row_id)].value
            choice_2 = ws[prep[2]+str(row_id)].value
            choice_3 = ws[prep[3]+str(row_id)].value
            choice_4 = ws[prep[4]+str(row_id)].value
            answer = ws[prep[5]+str(row_id)].value

            if choice_1 and choice_2 and choice_3 and choice_4 and answer:
                print(question_text, choice_1, choice_2, choice_3, choice_4)

                ChoiceQuestion.objects.get_or_create(question_text = question_text,
                                                    choice_1=choice_1, 
                                                    choice_2=choice_2, 
                                                    choice_3=choice_3, 
                                                    choice_4=choice_4, 
                                                    answer= str(answer))

            

