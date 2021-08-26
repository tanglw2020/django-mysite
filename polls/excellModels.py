import os
from django.db import models
from django.core.exceptions import ValidationError
from django.utils.translation import gettext_lazy as _
from django.utils.html import format_html, format_html_join

import re
import shutil
import zipfile
from pathlib import Path
from openpyxl import load_workbook
from openpyxl import cell

BASE_DIR = Path(__file__).resolve().parent.parent
MEDIA_ROOT  = BASE_DIR / 'media'

def find_2rd_item_in_maps(item, maps):
    for item1, item2 in maps:
        if item1 == item: 
            return item2

def validate_xlsx(value):
    extension = value.name.split('.')[-1]
    if extension != 'xlsx':
        raise ValidationError(
            _(value.name+'不是excel文件(.xlsx)'),
            params={'value': value.name},
        )

Standard_Color_Maps = [
    ('FFC00000','深红'),
    ('FFFF0000','红色'),
    ('FFFFC000','橙色'),
    ('FFFFFF00','黄色'),
    ('FF92D050','浅绿'),
    ('FF00B050','绿色'),
    ('FF00B0F0','浅蓝'),
    ('FF0070C0','蓝色'),
    ('FF002060','深蓝'),
    ('FF7030A0','紫色'),
]

## 套用套用表格格式
Table_Style_Maps = [
    ('TableStyleLight9', '"表样式浅色9"'),
    ('TableStyleLight12','"表样式浅色12"'),
    ('TableStyleMedium4','"表样式中等深浅4"'),
    ('TableStyleMedium8','"表样式中等深浅8"'),
    ('TableStyleDark3','"表样式深色3"'),
    ('TableStyleDark6','"表样式深色6"'),
]

## 图表类型
Chart_Type_Choice = [
    ('barChart', '柱形图'),
    ('lineChart', '折线图'),
    ('pieChart', '饼图'),
    ('areaChart', '面积图'),
    ('scatterChart', 'XY散点图'),
    ('radarChart', '雷达图'),
]

Sort_Type_Choice = [
    ('升序','升序'),
    ('降序','降序'),
]


## conditional_formatting
## ["name", "type", "operator", "bottom", "percent", "aboveAverage"]
# {
# "greaterThan":["cellIs", "greaterThan", "None", "None", "None"],
# "lessThan":["cellIs", "lessThan", "None", "None", "None"],
# "equal":["cellIs", "equal", "None", "None", "None"],
# "duplicateValues":["duplicateValues", "None", "None", "None", "None"],
# "top10":["top10", "None", "None", "None", "None"],
# "top10p":["top10", "None", "None", "True", "None"],
# "tail10":["top10", "None", "True", "None", "None"],
# "tail10p":["top10", "None", "True", "True", "None"],
# "aboveAverage":["aboveAverage", "None", "None", "None", "None"],
# "belowAverage":["aboveAverage", "None", "None", "None", "False"] 
# }
Conditional_Formatting_Type_Choice = [
    ('greaterThan','突出显示单元格规则大于'),
    ('lessThan','突出显示单元格规则小于'),
    ('equal','突出显示单元格规则等于'),
    ('duplicateValues','突出显示单元格规则重复值'),
    ('top10','最前最后规则前10项'),
    ('top10p','最前最后规则前10%'),
    ('tail10','最前最后规则后10项'),
    ('tail10p','最前最后规则后10%'),
    ('aboveAverage','最前最后规则高于平均值'),
    ('belowAverage','最前最后规则低于平均值'),
]

Conditional_Formatting_Coloring_Choice = [
    ('浅红填充色深红色文本','浅红填充色深红色文本'), 
    ('黄填充色深黄色文本','黄填充色深黄色文本'), 
    ('绿填充色深绿色文本','绿填充色深绿色文本'), 
    ('红色边框','红色边框'), 
]


def is_cell_range_legal(p):
    if p == '':
        return False, '不能为空'
        
    reg = re.compile('([A-Z])(\d+):([A-Z])(\d+)')
    result = reg.match(p)
    if result:
        p1,p2,p3,p4 = result.group(1), int(result.group(2)), result.group(3), int(result.group(4))
        if p1>p3 or p2>p4 or (p1==p3 and p2==p4):
            return False, '[起始地址] 应当小于 [结束地址]，比如A2:B5'
    else:
        return False, '地址格式 [大写字母][数字]:[大写字母][数字]，比如A2:B5'

    return True, ''


def clean_cell_range(position, errors_dict, target_key):
    is_legal, errors = is_cell_range_legal(position)
    if not is_legal:
        errors_dict[target_key] = _(errors)

def clean_empty_item(item, errors_dict, target_key):
    if item == '': errors_dict[target_key] = _('不能为空')

####################### Excel 题目 #########################
class ExcelQuestion(models.Model):
    # 控制 表项显示文字，默认按类名object（n）显示
    def __str__(self):
        return 'Excel操作题' + str(self.id) 

    class Meta:
        verbose_name_plural = 'Excel操作题'  
        verbose_name = 'Excel操作题' 

    # 包含小题数量
    def question_num(self):
        return 5
    question_num.short_description='小题数量'

    def base_path_(self):
        return os.path.join(MEDIA_ROOT, 'upload_xlsx', str(self.id))

    # 题目内容
    def question_content(self):
        pre_description = '打开工作薄文件EXCEL.xlsx:\n'
        result_list = []

        if self.rename_sheet_op:
            result_list.append('将工作表Sheet1重命名为"'+self.new_sheet_name+'".')

        if self.merge_cell_op:
            result_list.append('将数据区域'+self.merge_cell_position+'合并为一个单元格，内容水平居中.')

        if self.color_cell_op:
            result_list.append('将数据区域'+self.color_cell_position+ \
            '文字设置为标准'+find_2rd_item_in_maps(self.color_cell_font,Standard_Color_Maps) \
            +', 填充设置为标准'+find_2rd_item_in_maps(self.color_cell_filling, Standard_Color_Maps)+'.')

        if self.conditional_formatting_op:
            if self.conditional_formatting_type not in ['greaterThan', 'lessThan', 'equal']:
                result_list.append('将数据区域'+self.conditional_formatting_position+ \
                '按"'+find_2rd_item_in_maps(self.conditional_formatting_type,Conditional_Formatting_Type_Choice)+ \
                    '"设置为"'+self.conditional_formatting_coloring+'".')
            else:
                result_list.append('将数据区域'+self.conditional_formatting_position+ \
                '按"'+find_2rd_item_in_maps(self.conditional_formatting_type,Conditional_Formatting_Type_Choice)+ \
                   self.conditional_formatting_param+'"设置为"'+self.conditional_formatting_coloring+'".')
        
        if self.table_style_op:
            result_list.append('将数据区域'+self.table_style_position+ \
                '设置为套用表格格式'+find_2rd_item_in_maps(self.table_style_choice, Table_Style_Maps)+'.')

        if self.chart_op:
            result_list.append('选取数据列"'+self.chart_data_name+'"内容，建立"' \
                +find_2rd_item_in_maps(self.chart_type, Chart_Type_Choice)+'"，图表标题为"'+self.chart_tiltle \
                    +'"，将图表插入'+self.chart_position+'区域.')

        if self.sort_op:
            result_list.append('将工作表内容按主关键字"'+ \
                self.keyword_1+'"的'+self.sort_type_1+'次序和次关键字"'+ \
                self.keyword_2+'"的'+self.sort_type_2+'次序进行排序' \
                +'.')

        if self.formula_maxmin_op:
            result_list.append(self.formula_maxmin_description)

        if self.formula_sumavg_op:
            result_list.append(self.formula_sumavg_description)

        return format_html(pre_description) + \
                format_html("<ol>") + \
                format_html_join(
                '\n', '<li style="color:black;">{}</li>',
                ((x,) for x in result_list)
                ) \
                + format_html("</ol>")
    question_content.short_description = '题目内容'

    def file_path_(self):
        if self.upload_excel:
            answer_file_path = self.upload_excel.path
        else:
            answer_file_path = ''
        return answer_file_path
    file_path_.short_description = '文件地址'


    def score_(self, answer_file_path):
        result_list = []
        ## check the files
        if answer_file_path and os.path.exists(answer_file_path):
            wb = load_workbook(answer_file_path)
            ws = wb.active 

            if self.rename_sheet_op:
                result = 'rename_sheet::'
                for name in wb.sheetnames:
                    if name == self.new_sheet_name: 
                        result = result+name
                        break
                result_list.append(result)

            if self.merge_cell_op:
                result = 'merge_cell::'
                for cell in ws.merged_cells:
                    if str(cell) == self.merge_cell_position: 
                        result = result+self.merge_cell_position
                        break
                result_list.append(result)

            if self.color_cell_op:
                result = 'color_cell::'
                positions = self.color_cell_position.split(':')
                for cell in positions:
                    fgcolor = ws[cell].fill.fgColor
                    fontcolor = ws[cell].font.color
                    if fgcolor.rgb == self.color_cell_filling \
                        and fontcolor.rgb == self.color_cell_font:
                        result = result+cell+self.color_cell_filling+self.color_cell_font+'-'
                result_list.append(result)

            if self.conditional_formatting_op:
                result = 'conditional_formatting::'
                for condf in ws.conditional_formatting:
                    # print(condf.cells,)
                    rules = condf.rules
                    if str(condf.cells) != self.conditional_formatting_position:
                        continue

                    for rule in rules:  
                        # print(rule.type)
                        if self.conditional_formatting_type in ('greaterThan', 'lessThan', 'equal'):
                            if str(rule.operator) == self.conditional_formatting_type:
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                        elif self.conditional_formatting_type == 'top10':
                            if str(rule.type) == 'top10' \
                                and str(rule.bottom) == 'None' \
                                    and str(rule.percent)== 'None':
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                        elif self.conditional_formatting_type == 'top10p':
                            if str(rule.type) == 'top10' \
                                and str(rule.bottom) == 'None' \
                                    and str(rule.percent)== 'True':
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                        elif self.conditional_formatting_type == 'tail10':
                            if str(rule.type) == 'top10' \
                                and str(rule.bottom) == 'True' \
                                    and str(rule.percent)== 'None':
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                        elif self.conditional_formatting_type == 'tail10p':
                            if str(rule.type) == 'top10' \
                                and str(rule.bottom) == 'True' \
                                    and str(rule.percent)== 'True':
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                        elif self.conditional_formatting_type == 'duplicateValues':
                            if str(rule.type) == 'duplicateValues':
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                        elif self.conditional_formatting_type == 'aboveAverage':
                            if str(rule.type) == 'aboveAverage' \
                                and str(rule.aboveAverage)=="None":
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                        elif self.conditional_formatting_type == 'belowAverage':
                            if str(rule.type) == 'aboveAverage' \
                                and str(rule.aboveAverage)=="False":
                                result = result+self.conditional_formatting_position+'-'\
                                    +self.conditional_formatting_type+'-'+self.conditional_formatting_param
                                # break
                result_list.append(result)

            if self.table_style_op:
                result = 'table_style::'
                for table in ws._tables:
                    # print('套用表格格式:',table.name, table.ref, table.tableStyleInfo.name)
                    if str(table.ref) == self.table_style_position \
                        and table.tableStyleInfo.name == self.table_style_choice: 
                        result = result+self.table_style_position+'-'+self.table_style_choice
                result_list.append(result)

            if self.chart_op:
                result = 'chart::'
                for chart in ws._charts:
                    # print(chart.tagname, self.chart_type)
                    if chart.tagname == self.chart_type:
                        result = result+self.chart_type
                result_list.append(result)

            if self.sort_op:
                result = 'sort::'
                p1, p2 = self.sort_data_position_1.split(':')
                p3, p4 = self.sort_data_position_2.split(':')
                sort_result1 = [x.strip() for x in self.sort_data_result_1.split('\n')]
                sort_result2 = [x.strip() for x in self.sort_data_result_2.split('\n')]
                # print(ws[p1].value, ws[p2].value,ws[p3].value,ws[p4].value,)
                # print(sort_result1[0], sort_result1[-1], sort_result2[0], sort_result2[-1])
                if str(ws[p1].value)==sort_result1[0] and str(ws[p2].value)==sort_result1[-1] \
                    and str(ws[p3].value)==sort_result2[0] and str(ws[p4].value)==sort_result2[-1]:
                    result = result+self.keyword_1+'-'+self.sort_type_1
                result_list.append(result)



        else:
            result_list.append('Nothing to score')
        return result_list


    def test_(self):
        ## only for development
        if self.upload_excel:
            answer_file_path = self.upload_excel.path
        else:
            answer_file_path = ''
        return format_html("<ol>") + \
                format_html_join(
                '\n', '<li style="color:black;">{}</li>',
                ((x,) for x in self.score_(answer_file_path))
                ) \
                + format_html("</ol>")
    test_.short_description = '测试结果'

    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)  
        
    def clean(self):
        error_dict = {}

        if self.rename_sheet_op:
            clean_empty_item(self.new_sheet_name, error_dict, 'new_sheet_name')

        if self.merge_cell_op:
            clean_cell_range(self.merge_cell_position, error_dict, 'merge_cell_position')

        if self.color_cell_op:
            clean_cell_range(self.color_cell_position, error_dict, 'color_cell_position')

            if self.color_cell_font == '':
                error_dict['color_cell_font'] = _('必须选择一种颜色')
            if self.color_cell_filling == '':
                error_dict['color_cell_filling'] = _('必须选择一种颜色')
            if self.color_cell_font and self.color_cell_font==self.color_cell_filling:
                error_dict['color_cell_font'] = _('文字和填充不能用相同颜色')
                error_dict['color_cell_filling'] = _('文字和填充不能用相同颜色')

        if self.conditional_formatting_op:
            clean_empty_item(self.conditional_formatting_param, error_dict, 'conditional_formatting_param')
            clean_empty_item(self.conditional_formatting_type, error_dict, 'conditional_formatting_type')
            clean_empty_item(self.conditional_formatting_coloring, error_dict, 'conditional_formatting_coloring')
            clean_cell_range(self.conditional_formatting_position, error_dict, 'conditional_formatting_position')


        if self.table_style_op:
            clean_empty_item(self.table_style_choice, error_dict, 'table_style_choice')
            clean_cell_range(self.table_style_position, error_dict, 'table_style_position')


        if self.chart_op:
            clean_empty_item(self.chart_data_name, error_dict, 'chart_data_name')
            clean_empty_item(self.chart_type, error_dict, 'chart_type')
            clean_empty_item(self.chart_tiltle, error_dict, 'chart_tiltle')
            clean_cell_range(self.chart_data_position, error_dict, 'chart_data_position')
            clean_cell_range(self.chart_position, error_dict, 'chart_position')
        

        if self.sort_op:
            if self.keyword_1 == '': error_dict['keyword_1'] = _('不能为空')
            if self.sort_type_1 == '': error_dict['sort_type_1'] = _('不能为空')
            if self.sort_data_result_1 == '': error_dict['sort_data_result_1'] = _('不能为空')
            if self.keyword_2 == '': error_dict['keyword_2'] = _('不能为空')
            if self.sort_type_2 == '': error_dict['sort_type_2'] = _('不能为空')
            if self.sort_data_result_2 == '': error_dict['sort_data_result_2'] = _('不能为空')
            clean_cell_range(self.sort_data_position_1, error_dict, 'sort_data_position_1')
            clean_cell_range(self.sort_data_position_2, error_dict, 'sort_data_position_2')


        if self.formula_maxmin_op:
            clean_empty_item(self.formula_maxmin_description, error_dict, 'formula_maxmin_description')
            clean_empty_item(self.formula_maxmin_result_regex, error_dict, 'formula_maxmin_result_regex')
            clean_cell_range(self.formula_maxmin_data_position, error_dict, 'formula_maxmin_data_position')
            clean_cell_range(self.formula_maxmin_result_position, error_dict, 'formula_maxmin_result_position')

        if self.formula_sumavg_op:
            clean_empty_item(self.formula_sumavg_description, error_dict, 'formula_sumavg_description')
            clean_empty_item(self.formula_sumavg_result_regex, error_dict, 'formula_sumavg_result_regex')
            clean_cell_range(self.formula_sumavg_data_position, error_dict, 'formula_sumavg_data_position')
            clean_cell_range(self.formula_sumavg_result_position, error_dict, 'formula_sumavg_result_position')

        if self.upload_excel:
            print('')
        else:
            error_dict['upload_excel'] = _('必须上传文件')

        if len(error_dict)>0:
            raise ValidationError(error_dict)

    ## 
    upload_excel = models.FileField(verbose_name='上传EXCEL文件', upload_to='upload_xlsx/', 
        null=True, blank=True, validators=[validate_xlsx])

    # 重命名工作表Sheet1
    # 题目示例：将Sheet1重命名new_sheet_name
    rename_sheet_op = models.BooleanField('考查重命名工作表？', default=False)
    new_sheet_name = models.CharField('工作表新名称', max_length=50, blank=True, default='数据汇总表')

    # 合并单元格
    # 题目示例：将工作表Start:End单元格合并为一个单元格，内容水平居中
    merge_cell_op = models.BooleanField('考查合并单元格？', default=False)
    merge_cell_position = models.CharField('合并单元格区域[Start:End]', max_length=20, blank=True, default='A1:F1')

    # 设置单元格颜色和填充色
    # 题目示例：将工作表Start:End单元格文字设置成标准色X，填色设置成标准色Y
    color_cell_op = models.BooleanField('考查单元格颜色设置？', default=False)
    color_cell_position = models.CharField('单元格区域[Start:End]', max_length=20, blank=True, default='A2:A20')
    color_cell_font = models.CharField('文字颜色', choices=Standard_Color_Maps, max_length=20, blank=True, default='') 
    color_cell_filling = models.CharField('填充颜色', choices=Standard_Color_Maps, max_length=20, blank=True, default='') 


    # 设置单元格 条件格式
    # 题目示例：对工作表Start:End单元格区域应用条件格式
    conditional_formatting_op = models.BooleanField('考查单元格条件格式设置？', default=False)
    conditional_formatting_position = models.CharField('单元格区域[Start:End]', max_length=20, blank=True, default='A2:A20')
    conditional_formatting_type = models.CharField('条件类型', choices=Conditional_Formatting_Type_Choice, max_length=50, blank=True, default='')
    conditional_formatting_param = models.CharField('条件格式参数', max_length=10, blank=True, default='5')
    conditional_formatting_coloring = models.CharField('颜色选项', choices=Conditional_Formatting_Coloring_Choice, max_length=50, blank=True, default='')


    # 单元格 套用表格格式
    # 题目示例：将工作表Start:End单元格套用表格格式
    table_style_op = models.BooleanField('考查单元格套用表格格式？', default=False)
    table_style_position = models.CharField('单元格区域[Start:End]', max_length=20, blank=True, default='A2:A20')
    table_style_choice = models.CharField('套用表格格式', choices=Table_Style_Maps, max_length=20, blank=True, default='') 


    # 插入图表
    # 题目示例：为工作表数据列X建立曲线图表，
    # 图表标题为Y，将图表插入表格区域Z
    chart_op = models.BooleanField('考查插入图表？', default=False)
    chart_data_name = models.CharField('数据列名称', max_length=20, blank=True, default='平均升学率')
    chart_data_position = models.CharField('数据列位置', max_length=20, blank=True, default='F2:F50')
    chart_type = models.CharField('图表类型', choices=Chart_Type_Choice, max_length=20, blank=True, default='') 
    chart_tiltle = models.CharField('图表标题', max_length=20, blank=True, default='曲线图') 
    chart_position = models.CharField('图表插入区域[Start:End]', max_length=20, blank=True, default='G2:K20')

    # 双关键字排序
    # 题目示例：将表格按关键字1升序、关键字2降序排序
    sort_op = models.BooleanField('考查排序？', default=False)
    keyword_1 = models.CharField('主关键字', max_length=20, blank=True, default='')
    sort_type_1 = models.CharField('主关键字次序', choices=Sort_Type_Choice, max_length=20, blank=True, default='') 
    sort_data_position_1 = models.CharField('主关键字数据区域', max_length=20, blank=True, default='G3:G10')
    sort_data_result_1 = models.TextField('主关键字数据列最终排序结果', blank=True, default='')
    keyword_2 = models.CharField('次关键字', max_length=20, blank=True, default='')
    sort_type_2 = models.CharField('次关键字次序', choices=Sort_Type_Choice, max_length=20, blank=True, default='') 
    sort_data_position_2 = models.CharField('次关键字数据区域', max_length=20, blank=True, default='F3:F10')
    sort_data_result_2 = models.TextField('次关键字数据列最终排序结果', blank=True, default='')


    # 公式 max/min
    # 题目示例：用公式计算工作表[Start:End单元格/]的最大值/最小值，将计算公式写在X单元格
    formula_maxmin_op = models.BooleanField('考查公式max/min？', default=False)
    formula_maxmin_description = models.TextField('题目文字描述', 
        default='用公式计算工作表[Start:End单元格/]的最大值/最小值，将计算公式写在X单元格')
    formula_maxmin_data_position = models.CharField('数据区域[Start:End]', max_length=20, blank=True, default='A2:A20')
    formula_maxmin_result_position = models.CharField('公式填写在单元格', max_length=20, blank=True, default='') 
    formula_maxmin_result_regex = models.CharField('公式匹配模式', max_length=50, 
                blank=True, default='=MAX\(([A-Z])(\d+):([A-Z])(\d+)\)') 


    # 公式 sum/average
    # 题目示例：用公式计算工作表[Start:End单元格/]的总和/平均值，将计算公式写在X单元格
    formula_sumavg_op = models.BooleanField('考查公式sum/average？', default=False)
    formula_sumavg_description = models.TextField('题目文字描述', 
        default='用公式计算工作表[Start:End单元格/]的总和/平均值，将计算公式写在X单元格')
    formula_sumavg_data_position = models.CharField('数据区域[Start:End]', max_length=20, blank=True, default='A2:A20')
    formula_sumavg_result_position = models.CharField('公式填写在单元格', max_length=20, blank=True, default='') 
    formula_sumavg_result_regex = models.CharField('公式匹配模式', max_length=50, 
                blank=True, default='=SUM\(([A-Z])(\d+):([A-Z])(\d+)\)') 

                